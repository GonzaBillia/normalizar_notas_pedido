import pandas as pd
import csv, os, sys
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def read_file(filepath):
    """
    Lee archivos .dat, .txt o .csv con delimitadores como tabulación, coma, punto y coma, barra vertical o espacio.
    
    Parámetros:
        filepath (str): Ruta del archivo a procesar.

    Retorna:
        pandas.DataFrame: Contenido del archivo en un DataFrame si la lectura es exitosa.
                         En caso de error, retorna un DataFrame vacío.
    """
    try:
        # Primero intentamos leer usando tabulación, con encoding "latin1"
        df = pd.read_csv(filepath, delimiter="\t", quoting=csv.QUOTE_NONE, encoding="latin1", on_bad_lines='skip')
        
        # Si se lee como una sola columna, probablemente el delimitador es otro
        if df.shape[1] == 1:
            print(f"El archivo '{filepath}' no parece estar delimitado por tabulación. Detectando delimitador...")
            with open(filepath, 'r', encoding="latin1") as f:
                first_line = f.readline()
            delimiters = [",", ";", "|", " "]
            detected_delimiter = None
            for delim in delimiters:
                if delim in first_line:
                    detected_delimiter = delim
                    break
            if not detected_delimiter:
                print(f"Error: No se pudo determinar un delimitador válido en '{filepath}'.")
                return pd.DataFrame()
            # Leer con el delimitador detectado (por ejemplo, la coma)
            df = pd.read_csv(filepath, delimiter=detected_delimiter, quoting=csv.QUOTE_NONE, encoding="latin1", on_bad_lines='skip')
            print(f"Archivo '{filepath}' leído correctamente con delimitador detectado: '{detected_delimiter}'")
        
        # Limpiar los nombres de columna para evitar problemas con espacios en blanco
        if not df.empty:
            df.columns = df.columns.str.replace('"', '').str.strip()
        
        return df

    except Exception as e:
        print(f"Error al leer el archivo '{filepath}': {str(e)}")
        return pd.DataFrame()




def select_columns(df, selected_columns):
    """
    Selecciona columnas específicas de un DataFrame basado en índices.

    Parámetros:
        df (pd.DataFrame): DataFrame original.
        selected_columns (list): Índices de las columnas a extraer.

    Retorna:
        pd.DataFrame: DataFrame con solo las columnas seleccionadas.
    """
    try:
        if df.empty:
            raise ValueError("❌ Error: El DataFrame está vacío, no se pueden seleccionar columnas.")

        # Verificar si las columnas existen en el DataFrame
        max_index = df.shape[1] - 1  # Último índice válido
        for col in selected_columns:
            if col > max_index:
                raise IndexError(f"❌ Error: La columna '{col}' no existe en el DataFrame (máximo índice: {max_index}).")

        df_selected = df.iloc[:, selected_columns]
        return df_selected

    except Exception as e:
        raise RuntimeError(f"❌ Error al seleccionar columnas: {str(e)}")



def add_user_columns(df, provider, account):
    """
    Añade dos columnas con datos fijos (Proveedor y Cuenta) a un DataFrame.

    Parámetros:
        df (pd.DataFrame): DataFrame procesado.
        provider (str): Nombre del proveedor.
        account (str): Nombre de la cuenta.

    Retorna:
        pd.DataFrame: DataFrame con las nuevas columnas añadidas.
    """
    df["Proveedor"] = provider
    df["Cuenta"] = account
    return df

def merge_and_save(files_data, output_path):
    """
    Une múltiples DataFrames y los guarda en un archivo XLSX.
    Además, valida que la cantidad de filas concatenadas sea igual a la suma de filas de cada DataFrame.
    
    Parámetros:
        files_data (list): Lista de DataFrames a combinar.
        output_path (str): Ruta donde guardar el archivo.
    """
    try:
        # Calcular el total esperado de filas sumando la cantidad de filas de cada DataFrame.
        total_expected_rows = sum(len(df) for df in files_data)
        print(f"Total esperado de filas: {total_expected_rows}")

        # Concatenar los DataFrames.
        final_df = pd.concat(files_data, ignore_index=True)
        actual_rows = final_df.shape[0]
        print(f"Total de filas concatenadas: {actual_rows}")

        # Validar la cantidad de filas.
        if actual_rows != total_expected_rows:
            print(f"Advertencia: Se esperaban {total_expected_rows} filas, pero se obtuvieron {actual_rows}.")

        # Guardar el DataFrame concatenado en Excel.
        final_df.to_excel(output_path, index=False, sheet_name="Datos Normalizados")
        print(f"✅ Archivo guardado exitosamente en {output_path}")
    except Exception as e:
        print(f"❌ Error al guardar el archivo: {str(e)}")


def load_column_template_json(mapping):
    """
    Retorna el mapeo de columnas y la lista de columnas finales para el proveedor,
    sin leer un archivo JSON, usando un objeto predefinido.
    """
    
    # La lista de columnas finales se obtiene de los valores del mapping
    final_columns = list(mapping.values())
    return final_columns


def standardize_dataframe(df, column_mapping, final_columns):
    """
    Ordena y renombra las columnas de un DataFrame basado en un template externo.

    Parámetros:
        df (pd.DataFrame): DataFrame con los datos procesados.
        column_mapping (dict): Diccionario con nombres actuales como claves y nombres estándar como valores.
        final_columns (list): Lista con el orden final de las columnas.

    Retorna:
        pd.DataFrame: DataFrame con las columnas renombradas y ordenadas.
    """
    # Renombrar columnas según el diccionario de mapeo
    df = df.rename(columns=column_mapping)

    # Filtrar solo las columnas definidas en el template y ordenarlas
    df = df[final_columns]

    if "Precio Unitario" in df.columns:
        def convert_to_float(x):
            # Si es una cadena, eliminar el separador de miles y cambiar la coma decimal por punto
            if isinstance(x, str):
                x = x.replace('.', '').replace(',', '.')
            return float(x)
        
        df["Precio Unitario"] = df["Precio Unitario"].apply(convert_to_float).round(2)


    # 🛠️ Formatear "Fecha" a dd/mm/yyyy (eliminar la hora si existe)
    if "Fecha" in df.columns:
        df["Fecha"] = pd.to_datetime(df["Fecha"], errors='coerce').dt.strftime("%d/%m/%Y")


    return df


def style_excel_file(file_path):
    """
    Aplica estilos a un archivo Excel:
    - Fondo celeste claro para las cabeceras
    - Texto en negrita en las cabeceras
    - Ajuste automático del ancho de columna
    - Intercalado de colores en filas (blanco y celeste claro)
    - Bordes finos verticales entre columnas
    - Formato de código de barras como texto para evitar notación científica
    """
    try:
        # Cargar el archivo Excel
        wb = load_workbook(file_path)
        ws = wb.active  # Obtener la hoja activa

        # Definir estilos
        header_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")  # Celeste claro
        header_font = Font(bold=True, color="000000")  # Negrita, color negro
        row_fill = PatternFill(start_color="F2F8FC", end_color="F2F8FC", fill_type="solid")  # Fondo alterno para filas
        border_style = Side(border_style="thin", color="000000")  # Bordes finos

        # Aplicar estilos a las cabeceras
        for cell in ws[1]:  # Primera fila (cabecera)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Identificar el índice de la columna "Codigo de Barras"
        barcode_col_index = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value and str(cell.value).lower() == "codigo de barras":
                barcode_col_index = col_idx
                break

        # Aplicar estilos a las filas (intercalado de colores y bordes)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
            for cell in row:
                if row_idx % 2 == 0:  # Filas pares con fondo celeste claro
                    cell.fill = row_fill
                cell.border = Border(left=border_style, right=border_style)

                # Aplicar formato de texto al código de barras
                if barcode_col_index and cell.column == barcode_col_index:
                    cell.number_format = "@"  # Texto

        # Ajustar ancho de columnas automáticamente
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2  # Ajuste del ancho

        # Guardar los cambios en el archivo
        wb.save(file_path)
        print(f"🎨 Estilos aplicados correctamente al archivo: {file_path}")

        # 🛠️ Abrir la carpeta de destino automáticamente
        folder_path = os.path.dirname(file_path)
        open_folder(folder_path)

    except Exception as e:
        print(f"❌ Error al aplicar estilos al archivo Excel: {str(e)}")


def open_folder(folder_path):
    """
    Abre la carpeta en la que se guardó el archivo Excel.

    Parámetros:
        folder_path (str): Ruta de la carpeta que se abrirá.
    """
    try:
        if os.name == "nt":  # Windows
            os.startfile(folder_path)
        else:
            print("⚠️ No se pudo abrir la carpeta automáticamente.")
    except Exception as e:
        print(f"❌ Error al intentar abrir la carpeta: {str(e)}")


